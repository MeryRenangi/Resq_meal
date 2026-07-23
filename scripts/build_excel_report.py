#!/usr/bin/env python3
"""Convert JSON E2E test reports into Excel workbooks."""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1B5E20")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
PASS_FILL = PatternFill("solid", fgColor="E8F5E9")
FAIL_FILL = PatternFill("solid", fgColor="FFEBEE")
SUMMARY_FILL = PatternFill("solid", fgColor="F1F8E9")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _autosize_columns(ws, min_width: int = 10, max_width: int = 55) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value is not None:
                max_len = max(max_len, min(len(str(value)) + 2, max_width))
        ws.column_dimensions[letter].width = max_len


def _write_summary_sheet(wb: Workbook, reports: list[dict]) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "ResQ Meal — E2E Test Report Summary"
    ws["A1"].font = Font(bold=True, size=14, color="1B5E20")
    ws.merge_cells("A1:F1")

    ws["A3"] = "Generated"
    ws["B3"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    ws["A4"] = "Total Suites"
    ws["B4"] = len(reports)

    headers = ["Platform", "Runner", "Device", "Total", "Passed", "Failed", "Pass Rate"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=6, column=col, value=header)
    _style_header_row(ws, 6, len(headers))

    row = 7
    grand_total = grand_passed = grand_failed = 0
    for report in reports:
        ws.cell(row=row, column=1, value=report["platform"].title())
        ws.cell(row=row, column=2, value=report["runner"])
        ws.cell(row=row, column=3, value=report["device"])
        ws.cell(row=row, column=4, value=report["total_cases"])
        ws.cell(row=row, column=5, value=report["passed"])
        ws.cell(row=row, column=6, value=report["failed"])
        ws.cell(row=row, column=7, value=f"{report['pass_rate']}%")
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).fill = SUMMARY_FILL
        grand_total += report["total_cases"]
        grand_passed += report["passed"]
        grand_failed += report["failed"]
        row += 1

    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=4, value=grand_total)
    ws.cell(row=row, column=5, value=grand_passed)
    ws.cell(row=row, column=6, value=grand_failed)
    rate = round(grand_passed / grand_total * 100, 2) if grand_total else 0
    ws.cell(row=row, column=7, value=f"{rate}%")
    for col in range(1, 8):
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN_BORDER

    _autosize_columns(ws)


def _write_detail_sheet(wb: Workbook, report: dict) -> None:
    title = f"{report['platform'].title()} ({report['runner']})"
    ws = wb.create_sheet(title=title[:31])

    headers = ["Test ID", "Module", "Test Case", "Status", "Duration (ms)", "Platform"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(headers))

    for row_idx, case in enumerate(report["test_cases"], start=2):
        values = [
            case["id"],
            case["module"],
            case["name"],
            case["status"],
            case["duration_ms"],
            case["platform"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER
            if col == 4:
                cell.fill = PASS_FILL if value == "PASSED" else FAIL_FILL
                cell.font = Font(bold=True, color="2E7D32" if value == "PASSED" else "C62828")

    ws.freeze_panes = "A2"
    _autosize_columns(ws)


def build_excel(reports: list[dict], output: Path) -> None:
    wb = Workbook()
    _write_summary_sheet(wb, reports)
    for report in reports:
        _write_detail_sheet(wb, report)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="Build Excel report from JSON test results")
    parser.add_argument(
        "--inputs",
        nargs="+",
        type=Path,
        required=True,
        help="One or more JSON report files",
    )
    parser.add_argument(
        "--output",
        type=Path,
        required=True,
        help="Output .xlsx path",
    )
    args = parser.parse_args()

    reports = []
    for path in args.inputs:
        if not path.exists():
            print(f"Missing report: {path}", file=sys.stderr)
            return 1
        reports.append(json.loads(path.read_text(encoding="utf-8")))

    build_excel(reports, args.output)
    print(f"Excel report saved -> {args.output} ({len(reports)} suite(s))")
    return 0


if __name__ == "__main__":
    sys.exit(main())
